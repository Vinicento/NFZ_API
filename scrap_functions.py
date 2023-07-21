import pandas as pd
import time
from tqdm import tqdm
import json
import urllib.request
import math
import requests
import openpyxl
from multiprocessing import Pool
import traceback
import scrap_functions


class operating_func:
    @staticmethod
    def make_request(url):
        try:
            with urllib.request.urlopen(url) as response:
                html = response.read().decode()
            return html
        except urllib.error.HTTPError as e:
            if e.code == 429:
                # Retry after waiting for 1 second
                time.sleep(1)
                return operating_func.make_request(url)


class basic_searches:

    @staticmethod
    def kwota_kontraktu_dla_produktu(year_from, year_to, branches, produkt, usluga, path):

        years = [year for year in range(year_from, year_to + 1)]
        wyniki = {}
        for y in years:
            wyniki[y] = {}

        for year in years:
            for branch in branches:  # wchodzi w województwo
                link = f"https://api.nfz.gov.pl/app-umw-api/agreements?year={year}&branch={branch}&serviceType={usluga}&productCode={produkt}&page=1&limit=10&format=json&api-version=1.2"
                # link=f"https://api.nfz.gov.pl/app-umw-api/providers/{year}?productCode={produkt}&serviceType={usluga}&page=1&limit=25&format=json&api-version=1.2"
                response = requests.get(link)
                data = response.json()
                current = data["links"]['self']
                while (current != 'null') & (current != 'None') & (current != None) & (
                        "next" in data["links"]):  # wchodzi w szpitale
                    response = requests.get(current)
                    data = response.json()
                    current = data["links"]['next']
                    for i in range(len(data["data"]["agreements"])):
                        contract_link = data["data"]["agreements"][i]['links'][
                            "related"]  
                        cost = 0
                        while (contract_link != 'null') & (contract_link != 'None') & (
                                contract_link != None):  # wchodzi w kontrakty
                            contracts_response = requests.get(contract_link)
                            provider_contracts = contracts_response.json()
                            contract_link = provider_contracts["links"]['next']
                            for j in range(len(provider_contracts['data']['plans'])):
                                product = provider_contracts['data']['plans'][j]['attributes'][
                                    'product-code']  
                                if product == produkt:
                                    cost += provider_contracts['data']['plans'][j]['attributes']['price']
                            provider_name = provider_contracts['data']['agreement']['attributes']['provider-name']

                        if str(branch) in wyniki[year]:
                            new_row = pd.DataFrame({"nazwa": [provider_name], "kwota": [cost]})
                            wyniki[year][str(branch)] = pd.concat([wyniki[year][str(branch)], new_row])
                        else:
                            wyniki[year][str(branch)] = pd.DataFrame({"nazwa": [provider_name], "kwota": [cost]})

        writer = pd.ExcelWriter(path, engine='xlsxwriter')

        for year in years:
            sheetname = str(year)
            df = pd.concat(wyniki[year].values(), keys=wyniki[year].keys())
            df.to_excel(writer, sheet_name=sheetname)

        writer.save()

    ###############################################################
    @staticmethod
    def produkt_dla_szpitali(year_from, year_to, branches, path):

        years = [year for year in range(year_from, year_to + 1)]
        results = []
        for year in tqdm(years):
            for branch in tqdm(branches):
                first_main = f"https://api.nfz.gov.pl/app-umw-api/agreements?year={year}&branch={branch}&serviceType=03&page=1&limit=25&format=json"
                first_main_request = scrap_functions.operating_func.make_request(first_main)
                main_page = json.loads(first_main_request)
                pages = math.ceil(main_page['meta']["count"] / 25)

                for i in range(1, pages + 1):

                    first_main = f"https://api.nfz.gov.pl/app-umw-api/agreements?year={year}&branch={branch}&serviceType=03&page={i}&limit=25&format=json"
                    try:
                        main_page_request = scrap_functions.operating_func.make_request(first_main)
                        main_page = json.loads(main_page_request)
                    except:
                        pass
                    for i in main_page["data"]["agreements"]:
                        hospital_summary = i['attributes']['amount']
                        try:
                            minor_page_link = i['links']['related']
                        except:
                            pass
                        hospital_name = i['attributes']["provider-name"]

                        try:
                            minor_page_request = scrap_functions.operating_func.make_request(minor_page_link)
                            minor_page = json.loads(minor_page_request)
                        except:
                            pass
                        for i in minor_page['data']['plans']:
                            produkt = i['attributes']['product-name']
                            avg_price = i['attributes']['avg-price']
                            product_amount_summary = i['attributes']['price']
                            product_count_summary = i['attributes']['unit-count']
                            results.append([year, branch, hospital_name, produkt, avg_price, hospital_summary,
                                            product_amount_summary, product_count_summary])

        results = pd.DataFrame(results,
                               columns=["rok", "województwo", "szital", "produkt", "średnia_cena", "sumaryczna kwota",
                                        "Sumaryczna kwota kontraktu dla produktu", "Sumaryczna liczba  produktu"])

        writer = pd.ExcelWriter(path)
        for year in years:
            data_to_save = results[results['rok'] == year]
            data_to_save.to_excel(writer, sheet_name=f"year{year}", index=False)

        writer.save()

    @staticmethod
    def pacjenci_na_jgp(branches, list_of_codes, path):  # TODO rozdzielic zeby bylo wiadomo jaki kod ile

        full_data = {}
        for code in list_of_codes:
            link = f"https://api.nfz.gov.pl/app-stat-api-jgp/index-of-tables?catalog=1c&name={code}&format=json&api-version=1.1"
            response = requests.get(link)
            table_list = response.json()
            data_link_all_years = table_list['data']['attributes']['years']

            for i in range(len(data_link_all_years)):
                data_link = data_link_all_years[i]['tables']
                general_data = [table for table in data_link if table['type'] == 'general-data'][0]["id"]

                data_link = f"https://api.nfz.gov.pl/app-stat-api-jgp/basic-data/{general_data}?branch=true&page=1&limit=25&format=json&api-version=1.1"
                year_data = requests.get(data_link)
                year_data = year_data.json()

                year = year_data['data']['attributes']['year']

                data_years_total = year_data['data']['attributes']['data']
                for num, branch in enumerate(branches):
                    num_of_patients = data_years_total[num]["number-of-patients"]
                    if year not in full_data:
                        full_data[year] = {}
                    if branch not in full_data[year]:
                        full_data[year][branch] = 0
                    full_data[year][branch] += num_of_patients

                    # Create an Excel writer object
        writer = pd.ExcelWriter(path, engine='xlsxwriter')
        wojewodztwa = ['Dolnośląskie', 'Kujawsko-Pomorskie', 'Lubelskie', 'Lubuskie', 'Łódzkie', 'Małopolskie',
                       'Mazowieckie', 'Opolskie', 'Podkarpackie', 'Podlaskie', 'Pomorskie', 'Śląskie', 'Świętokrzyskie',
                       'Warmińsko-Mazurskie', 'Wielkopolskie', 'Zachodniopomorskie']
        frame = pd.DataFrame({"Woj": wojewodztwa})
        for year in list(full_data.keys()):
            frame[year] = full_data[year].values()
        frame.to_excel(writer, sheet_name="hospitalizacje")
        # Save the Excel file
        writer.save()

    @staticmethod
    def pakiet_kody(year_from, year_to, branches, products, path):

        years = range(year_from, year_to + 1)

        wyniki = {}
        for y in years:
            wyniki[y] = pd.DataFrame()

        for year in years:
            for search_product in tqdm(products):
                for branch in branches:  
                    link = f"https://api.nfz.gov.pl/app-umw-api/agreements?year={year}&branch={branch}&productCode={search_product}&page=1&limit=25&format=json&api-version=1.2"
                    response = requests.get(link)
                    data = response.json()
                    try:
                        pages = math.ceil(data['meta']["count"] / 25) + 1
                        for i in range(1, pages):
                            current = link = f"https://api.nfz.gov.pl/app-umw-api/agreements?year={year}&branch={branch}&productCode={search_product}&page={i}&limit=25&format=json&api-version=1.2"
                            response = requests.get(current)
                            data = response.json()

                            for j in range(len(data["data"]["agreements"])):
                                contract_link = data["data"]["agreements"][j]['links'][
                                    "related"]  
                                response = requests.get(contract_link)
                                data_2 = response.json()
                                try:
                                    pages_2 = math.ceil(data_2['meta']["count"] / 25)

                                    for i in range(1, pages_2 + 1):
                                        contract_link_2 = contract_link + "&page=" + str(i) + "&limit=25"
                                        contracts_response = requests.get(contract_link_2)
                                        provider_contracts = contracts_response.json()
                                        for j in range(len(provider_contracts['data']['plans'])):
                                            product = provider_contracts['data']['plans'][j]['attributes'][
                                                'product-code']  
                                            if product == search_product:
                                                cost = provider_contracts['data']['plans'][j]['attributes']['price']
                                                provider_name = provider_contracts['data']['agreement']['attributes'][
                                                    'provider-name']
                                                avg_price = provider_contracts['data']['plans'][j]['attributes'][
                                                    'avg-price']
                                                amount = provider_contracts['data']['plans'][j]['attributes'][
                                                    'unit-count']
                                                new_row = pd.DataFrame(
                                                    {"dostawca": [provider_name], "wojew": [branch], "kwota": [cost],
                                                     "ilosc": [amount], "kod produktu": [product],
                                                     "średnia cena ": [avg_price]})
                                                wyniki[year] = pd.concat([wyniki[year], new_row])
                                except:
                                    print(current)
                    except:
                        print(link)
        wojewodztwa = {
            '01': 'dolnośląskie',
            '02': 'kujawsko-pomorskie',
            '03': 'lubelskie',
            '04': 'lubuskie',
            '05': 'łódzkie',
            '06': 'małopolskie',
            '07': 'mazowieckie',
            '08': 'opolskie',
            '09': 'podkarpackie',
            '10': 'podlaskie',
            '11': 'pomorskie',
            '12': 'śląskie',
            '13': 'świętokrzyskie',
            '14': 'warmińsko-mazurskie',
            '15': 'wielkopolskie',
            '16': 'zachodniopomorskie'
        }


        writer = pd.ExcelWriter(path)
        for year in years:
            try:
                wyniki[year]['wojew'] = wyniki[year]['wojew'].map(wojewodztwa)

                data_to_save = wyniki[year]
                data_to_save.to_excel(writer, sheet_name=f"year{year}", index=False)
            except:
                pass
        writer.save()

    #######################################################################################

    @staticmethod
    def full_scrap(args):
        codes, icd_code = args
        link = "https://api.nfz.gov.pl/app-stat-api-jgp/benefits?catalog=1a&page=1&limit=25&format=json&api-version=1.1"
        # 2015:{},2016:{},2017:{},
        # Create an adapter with the retry strategy and connection pool
        final_data = {}
        year = 0
        full_list = {2016: {}, 2017: {}, 2018: {}, 2019: {}, 2020: {}}

        for code in tqdm(codes):

            link = f"https://api.nfz.gov.pl/app-stat-api-jgp/index-of-tables?catalog=1a&name={code}&format=json&api-version=1.1"

            try:
                html = operating_func.make_request(link)
                # wait for 1 second before making the next request

                table_list = json.loads(html)

            except:
                print(link)
            data_link_all_years = [year for year in table_list['data']['attributes']['years'] if year['year'] >= 2016]

            for i in range(len(data_link_all_years)):
                year = data_link_all_years[i]['year']

                data_link = data_link_all_years[i]['tables']

                if data_link:
                    general_data = [table for table in data_link if (table['type'] == 'icd-9-procedures') & (
                            table['attributes']['header'] == "Procedury ICD 9")][0]["id"]
                    full_list[year][code] = general_data

        for year in full_list.keys():
            for code in tqdm(full_list[year].keys()):
                general_data = full_list[year][code]

                current_icd = f"https://api.nfz.gov.pl/app-stat-api-jgp/icd9-procedures/{general_data}?format=json&limit=25&page=1"
                html_2 = operating_func.make_request(current_icd)
                data_icd = json.loads(html_2)
                pages = math.ceil(data_icd['meta']["count"] / 25)

                for i in range(1, pages + 1):
                    current_icd = f"https://api.nfz.gov.pl/app-stat-api-jgp/icd9-procedures/{general_data}?format=json&limit=25&page={i}"

                    try:

                        html_2 = operating_func.make_request(current_icd)
                        data_icd = json.loads(html_2)

                    except:
                        traceback.print_exc()
                        print(current_icd)

                    result = [table['number-of-hospitalizations'] for table in data_icd['data']["attributes"]["data"] if
                              table['procedure-code'] == str(icd_code)]
                    if result:
                        full_list[year][code] = result[0]

                        if year in final_data.keys():
                            final_data[year][code] = result
                        else:
                            final_data[year] = {code: result}

    @staticmethod
    def icd_9(icd_code):
        codes = []
        names = []

        scrap = requests.session()
        current = "https://api.nfz.gov.pl/app-stat-api-jgp/benefits?catalog=1a&page=1&limit=25&format=json&api-version=1.1"
        while (current != 'null') & (current != 'None') & (current != None):  # wchodzi w szpitale
            response = scrap.get(current)
            data = response.json()

            current = data["links"]['next']

            try:
                current = current[:4] + "s" + current[4:]
            except TypeError:
                pass

            for i in data['data']:
                codes.append(i['code'])
                names.append(i['name'])

        names = dict(zip(codes, names))


        with Pool(8) as p:

            results = p.map(basic_searches.full_scrap,
                            [(code_part, icd_code,) for code_part in [codes[i::8] for i in range(8)]])


        merged_dict = {}
        for d in [results[0], results[1], results[2], results[3], results[4], results[5], results[6], results[7]]:
            for year in d:
                if year not in merged_dict:
                    merged_dict[year] = {}
                merged_dict[year].update(d[year])

        new_results = {}

        for key, value in merged_dict.items():
            list_0f_vs = merged_dict[key].values()
            list_0f_vs = [x for x in list_0f_vs if isinstance(x, int)]
            filtered_dict = {k: v for k, v in merged_dict[key].items() if v in list_0f_vs}
            new_results[key] = filtered_dict


        workbook = openpyxl.Workbook()

        for year in new_results.keys():
            sheet = workbook.create_sheet(str(year))
            sheet['A1'] = 'Code'
            sheet['B1'] = 'Value'
            sheet['C1'] = 'Name'
            for i, code in enumerate(results[year].keys()):
                sheet.cell(row=i + 2, column=1, value=code)
                sheet.cell(row=i + 2, column=2, value=new_results[year][code])
                sheet.cell(row=i + 2, column=3, value=names[code])

        workbook.save('example.xlsx')

    
    @staticmethod
    def contract_by_service(*args):
        services = args[0][0]
        years = args[0][1]
        branches = args[0][2]
        results = []
        for service in services:
            for year in tqdm(years):
                for branch in tqdm(branches):
                    first_main = f"https://api.nfz.gov.pl/app-umw-api/agreements?year={year}&branch={branch}&serviceType={service}&page=1&limit=25&format=json"

                    first_main_request = operating_func.make_request(first_main)
                    if first_main_request is not None:
                        main_page = json.loads(first_main_request)
                        pages = math.ceil(main_page['meta']["count"] / 25)

                        for i in range(1, pages + 1):

                            first_main = f"https://api.nfz.gov.pl/app-umw-api/agreements?year={year}&branch={branch}&serviceType={service}&page={i}&limit=25&format=json"
                            try:
                                main_page_request = operating_func.make_request(first_main)
                                main_page = json.loads(main_page_request)
                            except:
                                try:
                                    time.sleep(1)
                                    main_page_request = operating_func.make_request(first_main)
                                    main_page = json.loads(main_page_request)
                                except:
                                    pass

                            for j in main_page["data"]["agreements"]:
                                hospital_summary = j['attributes']['amount']
                                try:
                                    minor_page_link = j['id']
                                except:
                                    pass
                                hospital_name = j['attributes']["provider-name"]

                                try:
                                    minor_page_request = operating_func.make_request(
                                        f"https://api.nfz.gov.pl/app-umw-api/agreements/{minor_page_link}?format=json&page=1&limit=25&api-version=1.2")
                                    minor_page = json.loads(minor_page_request)
                                except:
                                    try:
                                        time.sleep(1)
                                        minor_page_request = operating_func.make_request(
                                            f"https://api.nfz.gov.pl/app-umw-api/agreements/{minor_page_link}?format=json&page=1&limit=25&api-version=1.2")
                                        minor_page = json.loads(minor_page_request)
                                    except:
                                        pass

                                pages_minor = math.ceil(minor_page['meta']["count"] / 25)

                                for w in range(1, pages_minor + 1):
                                    try:
                                        minor_page_request = operating_func.make_request(
                                            f"https://api.nfz.gov.pl/app-umw-api/agreements/{minor_page_link}?format=json&page={w}&limit=25&api-version=1.2")
                                        minor_page = json.loads(minor_page_request)
                                    except:
                                        try:
                                            time.sleep(5)
                                            minor_page_request = operating_func.make_request(
                                                f"https://api.nfz.gov.pl/app-umw-api/agreements/{minor_page_link}?format=json&page={w}&limit=25&api-version=1.2")
                                            minor_page = json.loads(minor_page_request)
                                        except:
                                            pass
                                    for k in minor_page['data']['plans']:
                                        produkt = k['attributes']['product-name']
                                        avg_price = k['attributes']['avg-price']
                                        product_amount_summary = k['attributes']['price']
                                        product_count_summary = k['attributes']['unit-count']
                                        results.append(
                                            [year, service, branch, hospital_name, produkt, avg_price, hospital_summary,
                                             product_amount_summary, product_count_summary])

                    else:
                        pass
        return results

    @staticmethod
    def kwota_kontraktów(services, year_from, year_to, branches, path):

        # create the input list
        # create a multiprocessing pool with 4 processes
        years = [year for year in range(year_from, year_to + 1)]
        with Pool(8) as p:
            # split the input list into 4 equal parts and process each part in parallel
            # results = p.map(scraper.full_scrap, [codes[i::8] for i in range(8)])
            results = p.map(basic_searches.contract_by_service,
                            [(service, years, branches) for service in [services[i::8] for i in range(8)]])

        combined_df = pd.DataFrame()  # Initialize an empty DataFrame

        for sublist in results:
            df = pd.DataFrame(sublist, columns=["Rok", "Rodzaj świadczenia", "Województwo", "Nazwa świadczeniodawcy",
                                                "Nazwa produktu kontraktowanego", "Średnia cena produktu",
                                                "Kwota umowy", "Sumaryczna kwota kontraktu dla produktu",
                                                "Sumaryczna liczba kontraktu dla produktu"])
            combined_df = pd.concat([combined_df, df], ignore_index=True)

        writer = pd.ExcelWriter(path)
        for year in years:
            data_to_save = combined_df[combined_df['Rok'] == year]
            data_to_save.to_excel(writer, sheet_name=f"year{year}", index=False)

        writer.save()
